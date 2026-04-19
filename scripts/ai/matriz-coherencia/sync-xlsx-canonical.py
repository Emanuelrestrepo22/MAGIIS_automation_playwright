#!/usr/bin/env python3
"""
sync-xlsx-canonical.py - Sincroniza STRIPE_Test_Suite_Matriz_Sincronizado.xlsx
con la matriz .md canonica, eliminando el drift historico:

  1. Prefijo "E2E " en columna B (titulo).
  2. Formato corto "Hold y Cobro" / "sin Hold y Cobro" → expansion canonica
     "Hold desde Alta de Viaje y Cobro desde App Driver" / "sin Hold desde
     Alta de Viaje, Cobro desde App Driver".
  3. Artefacto "app modo personal" del rename previo.
  4. Doble "desde app pax ... desde app pax".

Alcance: columnas B (titulo) y C (descripcion) de TEST_SUITE y TEST_SUITE 2.0.
Lee matriz_cases.md para conseguir el titulo canonico por ID; si no existe
match exacto, aplica solo las sustituciones mecanicas (E2E, Hold y Cobro,
modo personal, doble desde app pax).

openpyxl preserva estilos y colores.
"""
from __future__ import annotations
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[3]
XLSX_PATH = ROOT / "docs" / "gateway-pg" / "stripe" / "STRIPE_Test_Suite_Matriz_Sincronizado.xlsx"
MD_PATHS = [
    ROOT / "docs" / "gateway-pg" / "stripe" / "matriz_cases.md",
    ROOT / "docs" / "gateway-pg" / "stripe" / "matriz_cases2.md",
]

# Acepta IDs TS-STRIPE-TCxxxx y TS-STRIPE-P2-TCxxx (y TS-STRIPE-TC-RVxxx)
ID_CORE = r"TS-STRIPE-(?:P\d+-)?TC[\w-]+"
TC_RE = re.compile(rf"\|\s*({ID_CORE})\s*\|\s*([^|]+?)\s*\|")
HEADER_PREFIX_RE = re.compile(rf"^{ID_CORE}\s*[\u2014\u2013\-]\s*", re.UNICODE)
TC_ID_PREFIX_RE = re.compile(rf"^({ID_CORE})\s*[\u2014\u2013\-]\s*(.+)$", re.UNICODE)


def load_canonical_titles() -> dict[str, str]:
    """Extrae de los md un mapa ID → titulo canonico (sin ** asteriscos, sin sufijo [NOTA])."""
    titles: dict[str, str] = {}
    for md_path in MD_PATHS:
        if not md_path.exists():
            continue
        for line in md_path.read_text(encoding="utf-8").splitlines():
            m = TC_RE.search(line)
            if not m:
                continue
            tc_id, desc = m.group(1).strip(), m.group(2).strip()
            if not tc_id.startswith("TS-STRIPE-"):
                continue
            # Saltar filas tipo "Alias de" (segunda columna es un ID)
            if desc.startswith("TS-STRIPE-"):
                continue
            # Limpiar markdown bold
            clean = desc.replace("**", "")
            # Quitar sufijos [NOTA...], (alias...), *(pendiente...*)
            clean = re.sub(r"\s*\[NOTA[^\]]*\]\s*", " ", clean, flags=re.IGNORECASE)
            clean = re.sub(r"\s*\(alias[^)]*\)", "", clean, flags=re.IGNORECASE)
            clean = re.sub(r"\s*\*\(pendiente[^)]*\)\*", "", clean, flags=re.IGNORECASE)
            clean = re.sub(r"\s+", " ", clean).strip()
            # Si arranca con "DEPRECADO" lo ignoramos (xlsx mantiene su texto propio)
            if clean.upper().startswith("DEPRECADO"):
                continue
            titles[tc_id] = clean
    return titles


def apply_mechanical_fixes(text: str) -> str:
    """Sustituciones seguras independientes del md:
       - "E2E " prefijo
       - "app modo personal" -> vacio (artefacto)
       - "Hold y Cobro" expansion canonica (positiva)
       - "sin Hold y Cobro" expansion canonica (sin hold)
       - "desde app pax ... desde app pax" -> collapsar a uno
    """
    s = text
    # 1. Eliminar prefijo "E2E " donde aparezca (tras "—" o al inicio)
    s = re.sub(r"\bE2E\s+", "", s)
    # 2. Limpieza "app modo personal" (artefacto del rename previo)
    #    "para usuario personal app modo personal" -> "para usuario personal"
    s = re.sub(r"para usuario personal\s+app modo personal", "para usuario personal", s)
    s = re.sub(r"\bapp modo personal\b", "", s)
    # 3. Expansion "sin Hold y Cobro" -> "sin Hold desde Alta de Viaje, Cobro desde App Driver"
    #    (hacer primero el "sin Hold" para no pisarlo con el general)
    s = re.sub(r"sin Hold y Cobro desde App Driver",
               "sin Hold desde Alta de Viaje, Cobro desde App Driver", s)
    # 4. Expansion "Hold y Cobro" -> "Hold desde Alta de Viaje y Cobro desde App Driver"
    s = re.sub(r"(?<!sin )Hold y Cobro desde App Driver",
               "Hold desde Alta de Viaje y Cobro desde App Driver", s)
    # 5. Colapsar "desde app pax ... desde app pax" (segunda ocurrencia dentro de misma oracion)
    s = re.sub(r"(desde app pax[^.]*?)\s+desde app pax\b", r"\1", s)
    # 6. Normalizar dobles espacios
    s = re.sub(r"\s{2,}", " ", s).strip()
    # 7. Limpiar punto final doble "..", espacios antes de puntuacion
    s = re.sub(r"\s+\.", ".", s)
    s = re.sub(r"\.{2,}", ".", s)
    return s


def main() -> int:
    canonical = load_canonical_titles()
    print(f"Titulos canonicos cargados desde md: {len(canonical)}")

    wb = openpyxl.load_workbook(XLSX_PATH)
    pre_e2e = 0
    pre_hyc = 0
    post_e2e = 0
    post_hyc = 0
    changed_rows = 0
    canonical_applied = 0
    mechanical_only = 0
    diffs: list[tuple[str, str, str, str]] = []  # (tc_id, col, old, new)

    for sn in ("TEST_SUITE", "TEST_SUITE 2.0"):
        ws = wb[sn]
        for row in range(2, ws.max_row + 1):
            b_val = ws.cell(row, 2).value
            c_val = ws.cell(row, 3).value
            if b_val is None:
                continue
            b_str = str(b_val)
            c_str = str(c_val) if c_val else ""

            # Counts pre
            if "E2E " in b_str: pre_e2e += 1
            if "E2E " in c_str: pre_e2e += 1
            if "Hold y Cobro" in b_str: pre_hyc += 1
            if "Hold y Cobro" in c_str: pre_hyc += 1

            m = TC_ID_PREFIX_RE.match(b_str)
            if not m:
                continue
            tc_id = m.group(1)

            if tc_id in canonical:
                new_title_body = canonical[tc_id]
                # reconstruir col B con el mismo separador que ya tenia (usar el que estaba)
                sep_match = re.search(rf"^{ID_CORE}\s*([\u2014\u2013\-])", b_str)
                sep = sep_match.group(1) if sep_match else "\u2014"
                new_b = f"{tc_id} {sep} {new_title_body}"
                new_c = new_title_body
                used_canonical = True
            else:
                # fallback: aplicar solo sustituciones mecanicas
                new_b = apply_mechanical_fixes(b_str)
                new_c = apply_mechanical_fixes(c_str) if c_str else c_str
                used_canonical = False

            row_changed = False
            if new_b != b_str:
                ws.cell(row, 2).value = new_b
                diffs.append((tc_id, "B", b_str, new_b))
                row_changed = True
            if new_c != c_str:
                ws.cell(row, 3).value = new_c
                diffs.append((tc_id, "C", c_str, new_c))
                row_changed = True
            if row_changed:
                changed_rows += 1
                if used_canonical:
                    canonical_applied += 1
                else:
                    mechanical_only += 1

            # Counts post
            if "E2E " in new_b: post_e2e += 1
            if "E2E " in new_c: post_e2e += 1
            if "Hold y Cobro" in new_b: post_hyc += 1
            if "Hold y Cobro" in new_c: post_hyc += 1

    wb.save(XLSX_PATH)

    print(f"\n=== Resumen ===")
    print(f"Filas modificadas: {changed_rows}")
    print(f"  con titulo canonico (md): {canonical_applied}")
    print(f"  solo sustituciones mecanicas: {mechanical_only}")
    print(f"\nOccurrences 'E2E '   pre={pre_e2e}  post={post_e2e}")
    print(f"Occurrences 'Hold y Cobro' (formato corto en celda sin contexto largo)")
    print(f"   nota: post puede tener 0 si expansion aplicada siempre")
    print(f"   pre={pre_hyc}  post={post_hyc}")

    print(f"\n=== Top 5 diffs ===")
    for tc_id, col, old, new in diffs[:5]:
        print(f"[{tc_id} col={col}]")
        print(f"  -- old: {old[:180]}")
        print(f"  ++ new: {new[:180]}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
