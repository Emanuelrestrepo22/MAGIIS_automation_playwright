#!/usr/bin/env python3
"""
sync-xlsx-rename-personal.py - Aplica al xlsx el rename semantico Eje 2
y los fixes TC082/083/088/089.

openpyxl preserva estilos/colores al abrir/guardar celdas sin tocar.
Alcance: columnas B (Titulo/ID) y C (Descripcion).
"""
from __future__ import annotations
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[3]
XLSX_PATH = ROOT / "docs" / "gateway-pg" / "stripe" / "STRIPE_Test_Suite_Matriz_Sincronizado.xlsx"


# Rename Eje 2: tipo de usuario "App Pax" -> "Personal" (preserva portal "App Pax")
RENAMES = [
    ("Usuario App Pax (Modo Personal)", "Usuario Personal"),
    ("Usuario App Pax (Modo Business / Colaborador)", "Usuario Business / Colaborador"),
    ("Usuario App Pax", "Usuario Personal"),
    ("usuario app pax modo personal", "usuario personal"),
    ("usuario app pax modo business", "usuario business"),
    ("usuario app pax existente", "usuario personal existente"),
    ("para usuario app pax con", "para usuario personal con"),
    ("para usuario app pax ", "para usuario personal "),
    ("para usuario app pax,", "para usuario personal,"),
    ("para usuario app pax.", "para usuario personal."),
    ("— usuario app pax", "— usuario personal"),
    (" usuario app pax ", " usuario personal "),
]

# Fixes Clonación -> Edición para TC082/083/088/089
TC_FIXES = {
    "TS-STRIPE-P2-TC082": [
        ("Clonación de viaje finalizado desde carrier", "Edición de viaje programado desde carrier"),
        ("Clonación de viaje finalizado para", "Edición de viaje programado para"),
    ],
    "TS-STRIPE-P2-TC083": [
        ("Clonación de viaje finalizado desde carrier", "Edición de viaje programado desde carrier"),
        ("Clonación de viaje finalizado para", "Edición de viaje programado para"),
    ],
    "TS-STRIPE-P2-TC088": [
        ("Clonación de viaje finalizado desde carrier", "Edición en conflicto desde carrier"),
        ("Clonación de viaje finalizado para", "Edición en conflicto para"),
    ],
    "TS-STRIPE-P2-TC089": [
        ("Clonación de viaje finalizado desde carrier", "Edición en conflicto desde carrier"),
        ("Clonación de viaje finalizado para", "Edición en conflicto para"),
    ],
}


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TEST_SUITE"]

    rename_changes = 0
    fix_changes = 0

    for row in range(2, ws.max_row + 1):
        b_cell = ws.cell(row, 2)
        c_cell = ws.cell(row, 3)
        b_val = str(b_cell.value) if b_cell.value is not None else ""
        c_val = str(c_cell.value) if c_cell.value is not None else ""

        # Detectar TC ID en col B
        matched_id = None
        for tid in TC_FIXES:
            if b_val.startswith(tid + " ") or b_val.startswith(tid + "\t") or b_val.startswith(tid + "|") or b_val.strip() == tid:
                matched_id = tid
                break

        # Fixes específicos por TC
        if matched_id:
            for old, new in TC_FIXES[matched_id]:
                if old in b_val:
                    new_b = b_val.replace(old, new)
                    b_cell.value = new_b
                    b_val = new_b
                    fix_changes += 1
                    print(f"  [fix] Row {row} {matched_id} col=B: {old!r} -> {new!r}")
                if old in c_val:
                    new_c = c_val.replace(old, new)
                    c_cell.value = new_c
                    c_val = new_c
                    fix_changes += 1
                    print(f"  [fix] Row {row} {matched_id} col=C: {old!r} -> {new!r}")

        # Renames Eje 2 generales (aplican a todas las filas)
        for old, new in RENAMES:
            if old in b_val:
                new_b = b_val.replace(old, new)
                b_cell.value = new_b
                b_val = new_b
                rename_changes += 1
            if old in c_val:
                new_c = c_val.replace(old, new)
                c_cell.value = new_c
                c_val = new_c
                rename_changes += 1

    wb.save(XLSX_PATH)
    print(f"\nRenames Eje 2 aplicados: {rename_changes}")
    print(f"Fixes TC082/083/088/089: {fix_changes}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
