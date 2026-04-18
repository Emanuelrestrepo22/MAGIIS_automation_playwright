#!/usr/bin/env python3
"""
sync-xlsx.py - Actualiza STRIPE_Test_Suite_Matriz_Sincronizado.xlsx aplicando
el fix de coherencia seccion<->descripcion en TC1011, TC1012, TC1016.

Alcance estricto: solo sustituye "Alta carrier de Viaje" -> "Alta de Viaje
desde app pax" en las columnas B (titulo) y C (descripcion) de esas 3 filas.
No toca: prefijo "E2E ", formato "Hold y Cobro" del xlsx, ni ninguna otra fila.

openpyxl preserva estilos y colores por defecto al abrir/guardar sin tocar
las celdas ajenas.
"""
from __future__ import annotations
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[3]
XLSX_PATH = ROOT / "docs" / "gateway-pg" / "stripe" / "STRIPE_Test_Suite_Matriz_Sincronizado.xlsx"

TARGET_IDS = {"TS-STRIPE-TC1011", "TS-STRIPE-TC1012", "TS-STRIPE-TC1016"}
# Pattern: reemplaza "Alta carrier de Viaje" por "Alta de Viaje desde app pax"
# case-insensitive para tolerar pequenas variaciones de capitalizacion.
PATTERN = re.compile(r"Alta\s+carrier\s+de\s+Viaje", re.IGNORECASE)
REPLACEMENT = "Alta de Viaje desde app pax"


def main() -> int:
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["TEST_SUITE"]
    changed = 0
    for row in range(2, ws.max_row + 1):
        b_val = ws.cell(row, 2).value  # Titulo (ID + descripcion)
        if not b_val:
            continue
        # Identificar TCs target SOLO si la celda B ARRANCA con el ID
        # (evita matches incidentales en notas "[NOTA: cubierta por TC1012]").
        b_str = str(b_val).strip()
        matched_id = None
        for tid in TARGET_IDS:
            if b_str.startswith(tid + " ") or b_str.startswith(tid + "\t") or re.match(rf"^{re.escape(tid)}\b", b_str):
                matched_id = tid
                break
        if not matched_id:
            continue
        # Aplicar sustitucion en columnas B (titulo) y C (descripcion)
        for col in (2, 3):
            cell = ws.cell(row, col)
            orig = cell.value
            if orig is None:
                continue
            new_val = PATTERN.sub(REPLACEMENT, str(orig))
            if new_val != orig:
                cell.value = new_val
                changed += 1
                col_label = {2: "B-titulo", 3: "C-desc"}[col]
                print(f"  [updated] Row {row} {matched_id} col={col_label}")
                print(f"    old: {orig[:140]}")
                print(f"    new: {new_val[:140]}")

    wb.save(XLSX_PATH)
    print(f"\nCeldas actualizadas: {changed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
